from os import link
from urllib.request import Request, urlopen
from bs4 import BeautifulSoup, SoupStrainer
from openpyxl.styles import Font
from openpyxl.utils import *
from openpyxl import *
import grequests
import datetime
import time

def main(nomorobo_page_index, pages_searched, coll):
    
    start = time.time()
    #delete certain links that we don't need
    def delete_things(links):
        del links[0:4]
        del links[-4:-1]
        del links[-1]
        if(nomorobo_page_index == '1'):
            links_to_delete = len(links)-120
            del links[0:links_to_delete]

    #Get webpages that contain URL's to detected robocalls
    req = Request("http://nomorobo.com/lookup?page=" + nomorobo_page_index) 
    html_page = urlopen(req)
    soup = BeautifulSoup(html_page, "lxml")

    #loop through the site and append all URL's into a list
    links = []
    def get_urls():
        for link in soup.findAll('a'):
            if link['href'].startswith('/'):
                link['href'] = 'http://nomorobo.com' + link['href']
            links.append(link.get('href'))
  
    get_urls()
    delete_things(links)

    gudnumbers = []

    #write found numbers to an Excel file, this looks a lot better than writing everything to a notepad
    def excel(coll, allnumbers):
        wb = load_workbook('Numbers.xlsx')
        sheet = wb["Numbers"]
        newcell = sheet.cell(len(allnumbers)+1, coll)
        newcell.value = allnumbers[-1]
        wb.save("Numbers.xlsx")

    #append the links containing the correct keywords to an array
    def assignn(x, links):
        if links[x] not in allnumbers:
            gudnumbers.append(links[x])
            allnumbers.append(links[x])
            print("After searching", (links.index(links[x]) + 1) , "numbers, I found: ", links[x])
            
            #write to a textfile, not the best option since its not readable whatsoever
            #textfile = open("Numbers.txt", "a")
            #textfile.write(links[x] + "\n")
            #textfile.close()

            excel(coll, allnumbers)
  
    #loop through the links and find keywords, this will decide if the link will be added to the list of refund/amazon scammers
    x = 0
    only_i_tags = SoupStrainer("i")
    spans = []
    gudwords = ['amazon', 'Amazon', 'computer', 'refund', '299', '399', 'Norton', 'Computer', 'Maccaffee', 'Refund']
    badwords = ['399-', '299-', '2999', '3999', 'Alexa', 'list', 'blood', 'Cross', 'cross', 'Congress', 'congress', 'health', 'Health']

    #Scraping all the found links to phonenumbers for certain keywords in certain HTML tags using multiprocessing
    req2 = (grequests.get(link) for link in links)
    pool = grequests.Pool(8)
    resp = grequests.imap(req2, pool)
    for r in resp: 
        moresoup = BeautifulSoup(r.text, "lxml", parse_only=only_i_tags)
        spans.append(moresoup)
        desc = str(spans[-1])
        numberstring = desc.split(' ')
        gudstring = any(item in gudwords for item in numberstring)
        if (bool(gudstring) == True):
            badstring = any(item in badwords for item in numberstring)
            if (bool(badstring) == False):
                assignn(links.index(x), links)
        pool.join()
        x += 1
        
    print("I have searched", pages_searched ,"page(s) so far")
    print("Numbers I found on the last page: ", gudnumbers)
    end = time.time()
    print("That took ", (end - start), "seconds")
  
allnumbers = []

#kind of the main function, user interface
def othermain(pages_to_search):

    #Assign some variables that will be used later in main()
    time.sleep(2)
    max_pages = 51
    pages_searched = 0
    nomorobo_page_index = '0'
    mainloops_done = 1
    coll = 1
    now = str(datetime.datetime.now())

    #More textfile stuff
    #textfile = open("Numbers.txt", "a")
    #textfile.write("\n" + "Date generated: " + now + "\n")
    #textfile.close()

    #Write the current date to the Excel file and prepare a new collumn to be filled 
    wb = load_workbook('Numbers.xlsx')
    sheet = wb["Numbers"]
    newcell = sheet.cell(len(allnumbers) + 2, column = coll)
    while (newcell.value):
        coll += 1
        newcell = sheet.cell(len(allnumbers) + 2, column = coll)
    datecell = sheet.cell(1, coll)
    datecell.value = now
    datecell.font = Font(bold=True)
    wb.save("Numbers.xlsx")
    
    #Loop through the chosen amount of Nomorobo webpages and scrape them for numbers using main(), only if the user chose a value from 0 to 51
    if(0 < pages_to_search < max_pages):
        print("Searching the most recent", pages_to_search , "Nomorobo page(s)...")
        for mainloops_done in range(pages_to_search):
            nomorobo_page_index = str(mainloops_done)
            mainloops_done += 1
            pages_searched += 1
            main(nomorobo_page_index, pages_searched, coll)

    #print out all found numbers in a nice list in the shell
    def printer():
        if(len(allnumbers) != 0):
            l = 1
            for number in allnumbers:
                print(l,":", number)
                l += 1
        else:
            print("We found nothing, sorry for wasting your time I guess")
    
    # log the amount of found numbers
    def chart():
        wb = load_workbook('Numbers.xlsx')
        sheet = wb["chart"]
        chart_coll = 1
        chart_row = 2
        active_cell = sheet.cell(chart_row, chart_coll)
        while(active_cell.value):
            chart_row += 1
            active_cell = sheet.cell(chart_row, chart_coll)
        badtime = str(datetime.datetime.now())
        active_cell.value = badtime[:len(badtime) - 10]
        active_cell = sheet.cell(chart_row, chart_coll + 1)
        active_cell.value = int(len(allnumbers))
        wb.save("Numbers.xlsx")

    #finish up by showing the results of the scraping, only if the user chose a value from 0 to 51
    if (0 < pages_to_search < 51):
        print("Final results: ")
        printer()
        chart()
        input()
    else:
        print("Please pick a value from 1 to 50.")
        othermain()

othermain(50)

